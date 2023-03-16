from bs4 import BeautifulSoup
import requests
import openpyxl
excle = openpyxl.Workbook()
# print(excle.sheetnames)
sheet = excle.active
sheet.title = "Top Rated Movies in IMDB"
# print(excle.sheetnames)
sheet.append(['Rank', 'Name', 'Year Of Release', 'rating'])
# # Url1 = requests.patch('https://www.naukri.com/jobs-in-kolkata')
# # print(Url1.text)
url = 'https://www.imdb.com/chart/top/'
# url = 'https://www.imdb.com'
try:
    response = requests.get(url)
    response.raise_for_status()
    # print(response.text)
    soup = BeautifulSoup(response.text, 'html.parser')

    movies = soup.find('tbody', class_="lister-list").find_all('tr')

    for movie in movies:
        rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]
        name = movie.find('td', class_="titleColumn").a.text
        year = movie.find('span', class_='secondaryInfo').text.strip('()')
        rating = movie.find('td', class_='ratingColumn imdbRating').strong.text
        print(rank, name, year, rating)
        sheet.append([rank,name,year,rating])




except Exception as e:
    print(e)
# # response = requests.get(url)
# # response.raise_for_status()
excle.save('IMDB Movie Rating.xlsx')

